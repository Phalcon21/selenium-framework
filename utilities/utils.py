import csv
import inspect
import logging
import softest
from openpyxl import load_workbook


class Utils(softest.TestCase):

    def custom_logger(logLevel: str = logging.DEBUG):
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("reports/automation.log")
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name: str, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename: str):
        datalist = []
        csvdata = open(filename, "r")
        reader = csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist
